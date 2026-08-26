#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
from dataclasses import dataclass
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from scipy.integrate import solve_ivp
from scipy.interpolate import PchipInterpolator
from scipy.optimize import brentq


MODEL_ID = "WHUT13-Q1-THERMAL"
MODEL_VERSION = "4438986"


@dataclass(frozen=True)
class ModelParameters:
    d: np.ndarray
    rho: np.ndarray
    c: np.ndarray
    k: np.ndarray
    body_mass: float = 60.0
    body_specific_heat: float = 3470.0
    body_area: float = 1.6521
    h_body: float = 3.0
    ambient_temperature: float = -40.0
    initial_temperature: float = 37.0

    @property
    def body_capacity(self) -> float:
        return self.body_mass * self.body_specific_heat


PARAMS = ModelParameters(
    d=np.array([0.0007, 0.0004, 0.0003], dtype=float),
    rho=np.array([208.0, 552.3, 300.0], dtype=float),
    c=np.array([4803.8, 2400.0, 5463.2], dtype=float),
    k=np.array([0.068, 0.06, 0.0527], dtype=float),
)


@dataclass
class ModalBasis:
    length: float
    lambdas: np.ndarray
    gamma: np.ndarray
    quad_y: np.ndarray
    quad_w: np.ndarray
    phi_quad: np.ndarray
    norms: np.ndarray
    int0: np.ndarray
    int1: np.ndarray
    phi_left: np.ndarray
    phi_right: np.ndarray
    dphi_left: np.ndarray
    dphi_right: np.ndarray


@dataclass
class ModalState:
    time_s: float
    body_temperature: float
    interface_12: float
    interface_23: float
    b1: np.ndarray
    b2: np.ndarray
    b3: np.ndarray
    basis3: ModalBasis
    h_external: float


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def load_pcm_curve(path: Path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["Sheet1"]
    temperatures = []
    raw_dsc = []
    for temperature, value in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        if temperature is None or value is None:
            continue
        temperatures.append(float(temperature))
        raw_dsc.append(float(value))
    temperatures = np.asarray(temperatures, dtype=float)
    raw_dsc = np.asarray(raw_dsc, dtype=float)
    if len(temperatures) < 2 or not np.all(np.diff(temperatures) > 0):
        raise ValueError("DSC temperatures must be strictly increasing")
    if not np.all(raw_dsc < 0):
        raise ValueError("Attachment header says exothermic is negative, but non-negative DSC values were found")
    specific_power = -1000.0 * raw_dsc
    interpolator = PchipInterpolator(temperatures, specific_power, extrapolate=False)
    lower = float(temperatures[0])
    upper = float(temperatures[-1])

    def q_pcm(temperature):
        values = np.asarray(temperature, dtype=float)
        inside = (values >= lower) & (values <= upper)
        result = np.zeros_like(values, dtype=float)
        if np.any(inside):
            result[inside] = interpolator(values[inside])
        if result.ndim == 0:
            return float(result)
        return result

    metadata = {
        "rows": int(len(temperatures)),
        "temperature_min_C": lower,
        "temperature_max_C": upper,
        "specific_power_min_W_per_kg": float(specific_power.min()),
        "specific_power_max_W_per_kg": float(specific_power.max()),
        "peak_temperature_C": float(temperatures[np.argmax(specific_power)]),
        "input_sha256": sha256(path),
    }
    return q_pcm, metadata


def gauss_grid(length: float, order: int):
    nodes, weights = np.polynomial.legendre.leggauss(order)
    return 0.5 * length * (nodes + 1.0), 0.5 * length * weights


def robin_roots(length: float, conductivity: float, coefficient: float, count: int):
    biot = coefficient * length / conductivity
    if biot <= 0:
        raise ValueError("Robin coefficient must be positive")
    roots = np.empty(count, dtype=float)
    epsilon = 1e-10
    for index in range(1, count + 1):
        left = (index - 0.5) * math.pi + epsilon
        right = index * math.pi - epsilon
        roots[index - 1] = brentq(lambda z: math.tan(z) + z / biot, left, right)
    return roots / length


def finalize_basis(length, lambdas, diffusivity, quad_y, quad_w, phi_quad, dphi_quad_fn):
    norms = np.sum(phi_quad * phi_quad * quad_w, axis=1)
    int0 = phi_quad @ quad_w
    int1 = phi_quad @ (quad_w * quad_y)
    phi_left = phi_quad[:, 0] * 0.0 + dphi_quad_fn("phi_left")
    phi_right = phi_quad[:, 0] * 0.0 + dphi_quad_fn("phi_right")
    dphi_left = phi_quad[:, 0] * 0.0 + dphi_quad_fn("dphi_left")
    dphi_right = phi_quad[:, 0] * 0.0 + dphi_quad_fn("dphi_right")
    return ModalBasis(
        length=length,
        lambdas=lambdas,
        gamma=diffusivity * lambdas * lambdas,
        quad_y=quad_y,
        quad_w=quad_w,
        phi_quad=phi_quad,
        norms=norms,
        int0=int0,
        int1=int1,
        phi_left=phi_left,
        phi_right=phi_right,
        dphi_left=dphi_left,
        dphi_right=dphi_right,
    )


def make_layer1_basis(params: ModelParameters, modes: int, quadrature: int):
    length = params.d[0]
    conductivity = params.k[0]
    coefficient = params.h_body
    diffusivity = conductivity / (params.rho[0] * params.c[0])
    lambdas = robin_roots(length, conductivity, coefficient, modes)
    quad_y, quad_w = gauss_grid(length, quadrature)
    ratio = coefficient / (conductivity * lambdas)
    phi_quad = np.cos(lambdas[:, None] * quad_y) + ratio[:, None] * np.sin(
        lambdas[:, None] * quad_y
    )

    def endpoint(kind):
        if kind == "phi_left":
            return np.ones(modes)
        if kind == "phi_right":
            return np.cos(lambdas * length) + ratio * np.sin(lambdas * length)
        if kind == "dphi_left":
            return np.full(modes, coefficient / conductivity)
        return -lambdas * np.sin(lambdas * length) + coefficient / conductivity * np.cos(
            lambdas * length
        )

    return finalize_basis(length, lambdas, diffusivity, quad_y, quad_w, phi_quad, endpoint)


def make_layer2_basis(params: ModelParameters, modes: int, quadrature: int):
    length = params.d[1]
    diffusivity = params.k[1] / (params.rho[1] * params.c[1])
    lambdas = np.arange(1, modes + 1, dtype=float) * math.pi / length
    quad_y, quad_w = gauss_grid(length, quadrature)
    phi_quad = np.sin(lambdas[:, None] * quad_y)

    def endpoint(kind):
        if kind in {"phi_left", "phi_right"}:
            return np.zeros(modes)
        if kind == "dphi_left":
            return lambdas
        return lambdas * np.cos(lambdas * length)

    return finalize_basis(length, lambdas, diffusivity, quad_y, quad_w, phi_quad, endpoint)


def make_layer3_basis(params: ModelParameters, h_external: float, modes: int, quadrature: int):
    length = params.d[2]
    conductivity = params.k[2]
    diffusivity = conductivity / (params.rho[2] * params.c[2])
    lambdas = robin_roots(length, conductivity, h_external, modes)
    quad_y, quad_w = gauss_grid(length, quadrature)
    phi_quad = np.sin(lambdas[:, None] * quad_y)

    def endpoint(kind):
        if kind == "phi_left":
            return np.zeros(modes)
        if kind == "phi_right":
            return np.sin(lambdas * length)
        if kind == "dphi_left":
            return lambdas
        return lambdas * np.cos(lambdas * length)

    return finalize_basis(length, lambdas, diffusivity, quad_y, quad_w, phi_quad, endpoint)


def external_coefficient(surface_temperature: float, params: ModelParameters):
    difference = surface_temperature - params.ambient_temperature
    if difference < -1e-8:
        raise ValueError("Outer surface fell below ambient temperature; check the numerical solution")
    return 2.38 * max(difference, 0.0) ** 0.25


def w1_values(y, body_temperature, interface_temperature, params: ModelParameters):
    denominator = params.k[0] + params.h_body * params.d[0]
    slope = params.h_body * (interface_temperature - body_temperature) / denominator
    intercept = (
        params.k[0] * interface_temperature
        + params.h_body * params.d[0] * body_temperature
    ) / denominator
    return intercept + slope * np.asarray(y)


def w2_values(y, interface_12, interface_23, params: ModelParameters):
    fraction = np.asarray(y) / params.d[1]
    return (1.0 - fraction) * interface_12 + fraction * interface_23


def w3_values(y, interface_temperature, h_external, params: ModelParameters):
    denominator = params.k[2] + h_external * params.d[2]
    slope = h_external * (params.ambient_temperature - interface_temperature) / denominator
    return interface_temperature + slope * np.asarray(y)


def evaluate_series(basis: ModalBasis, coefficients: np.ndarray, y):
    y_values = np.asarray(y, dtype=float)
    if basis.phi_left[0] == 1.0:
        ratio = PARAMS.h_body / (PARAMS.k[0] * basis.lambdas)
        phi = np.cos(basis.lambdas[:, None] * y_values.ravel()) + ratio[:, None] * np.sin(
            basis.lambdas[:, None] * y_values.ravel()
        )
    else:
        phi = np.sin(basis.lambdas[:, None] * y_values.ravel())
    values = coefficients @ phi
    return values.reshape(y_values.shape)


def project_values(values: np.ndarray, basis: ModalBasis):
    return (basis.phi_quad @ (basis.quad_w * values)) / basis.norms


def layer3_surface(state: ModalState, params: ModelParameters):
    base = w3_values(params.d[2], state.interface_23, state.h_external, params)
    return float(base + state.b3 @ state.basis3.phi_right)


def reproject_layer3(
    state: ModalState,
    new_basis: ModalBasis,
    new_h_external: float,
    params: ModelParameters,
):
    y = new_basis.quad_y
    old_temperature = w3_values(y, state.interface_23, state.h_external, params) + evaluate_series(
        state.basis3, state.b3, y
    )
    new_base = w3_values(y, state.interface_23, new_h_external, params)
    return project_values(old_temperature - new_base, new_basis)


def modal_layer_integrals(state: ModalState, basis1, basis2, params: ModelParameters):
    int1 = float(
        np.sum(
            basis1.quad_w
            * (
                w1_values(
                    basis1.quad_y, state.body_temperature, state.interface_12, params
                )
                + state.b1 @ basis1.phi_quad
            )
        )
    )
    int2 = float(
        np.sum(
            basis2.quad_w
            * (
                w2_values(
                    basis2.quad_y, state.interface_12, state.interface_23, params
                )
                + state.b2 @ basis2.phi_quad
            )
        )
    )
    int3 = float(
        np.sum(
            state.basis3.quad_w
            * (
                w3_values(
                    state.basis3.quad_y,
                    state.interface_23,
                    state.h_external,
                    params,
                )
                + state.b3 @ state.basis3.phi_quad
            )
        )
    )
    return np.array([int1, int2, int3])


def modal_total_energy(state, basis1, basis2, params):
    layer_integrals = modal_layer_integrals(state, basis1, basis2, params)
    clothing = params.body_area * np.sum(params.rho * params.c * layer_integrals)
    return params.body_capacity * state.body_temperature + clothing


def modal_average_layer2(state: ModalState, basis2: ModalBasis, params: ModelParameters):
    base_average = 0.5 * (state.interface_12 + state.interface_23)
    modal_average = float(state.b2 @ basis2.int0 / params.d[1])
    return base_average + modal_average


def modal_step(
    state: ModalState,
    dt: float,
    basis1: ModalBasis,
    basis2: ModalBasis,
    q_pcm,
    modes: int,
    quadrature: int,
    params: ModelParameters,
):
    surface_start = layer3_surface(state, params)
    h_external = external_coefficient(surface_start, params)
    basis3 = make_layer3_basis(params, h_external, modes, quadrature)
    b3_start = reproject_layer3(state, basis3, h_external, params)
    layer2_average = modal_average_layer2(state, basis2, params)
    pcm_specific_power = q_pcm(layer2_average)
    source_rate = pcm_specific_power / params.c[1]

    exponential1 = np.exp(-basis1.gamma * dt)
    exponential2 = np.exp(-basis2.gamma * dt)
    exponential3 = np.exp(-basis3.gamma * dt)
    gain1 = -np.expm1(-basis1.gamma * dt) / basis1.gamma
    gain2 = -np.expm1(-basis2.gamma * dt) / basis2.gamma
    gain3 = -np.expm1(-basis3.gamma * dt) / basis3.gamma

    p10 = basis1.int0 / basis1.norms
    p11 = basis1.int1 / basis1.norms
    p20 = basis2.int0 / basis2.norms
    p21 = basis2.int1 / basis2.norms
    p30 = basis3.int0 / basis3.norms
    p31 = basis3.int1 / basis3.norms

    def evaluate(beta, return_state=False):
        beta_h, beta_12, beta_23 = beta
        denominator1 = params.k[0] + params.h_body * params.d[0]
        a1_dot = (
            params.k[0] * beta_12 + params.h_body * params.d[0] * beta_h
        ) / denominator1
        b1_dot = params.h_body * (beta_12 - beta_h) / denominator1
        force1 = -(a1_dot * p10 + b1_dot * p11)

        force2 = (
            source_rate * p20
            - beta_12 * (p20 - p21 / params.d[1])
            - beta_23 * p21 / params.d[1]
        )

        denominator3 = params.k[2] + h_external * params.d[2]
        force3 = -beta_23 * (p30 - h_external * p31 / denominator3)

        b1_end = state.b1 * exponential1 + force1 * gain1
        b2_end = state.b2 * exponential2 + force2 * gain2
        b3_end = b3_start * exponential3 + force3 * gain3

        body_end = state.body_temperature + beta_h * dt
        interface_12_end = state.interface_12 + beta_12 * dt
        interface_23_end = state.interface_23 + beta_23 * dt

        slope1 = params.h_body * (interface_12_end - body_end) / denominator1
        slope2 = (interface_23_end - interface_12_end) / params.d[1]
        slope3 = h_external * (
            params.ambient_temperature - interface_23_end
        ) / denominator3
        derivative1_right = slope1 + b1_end @ basis1.dphi_right
        derivative2_left = slope2 + b2_end @ basis2.dphi_left
        derivative2_right = slope2 + b2_end @ basis2.dphi_right
        derivative3_left = slope3 + b3_end @ basis3.dphi_left

        residual1 = params.k[0] * derivative1_right - params.k[1] * derivative2_left
        residual2 = params.k[1] * derivative2_right - params.k[2] * derivative3_left

        a1_start = (
            params.k[0] * state.interface_12
            + params.h_body * params.d[0] * state.body_temperature
        ) / denominator1
        integral_body = state.body_temperature * dt + 0.5 * beta_h * dt * dt
        integral_w1_surface = a1_start * dt + 0.5 * a1_dot * dt * dt
        integral_b1 = state.b1 * gain1 + force1 / basis1.gamma * (dt - gain1)
        integral_t1_surface = integral_w1_surface + float(np.sum(integral_b1))
        residual3 = (
            params.body_capacity * beta_h * dt
            + params.body_area
            * params.h_body
            * (integral_body - integral_t1_surface)
        )
        residual = np.array([residual1, residual2, residual3], dtype=float)
        if return_state:
            return residual, ModalState(
                time_s=state.time_s + dt,
                body_temperature=body_end,
                interface_12=interface_12_end,
                interface_23=interface_23_end,
                b1=b1_end,
                b2=b2_end,
                b3=b3_end,
                basis3=basis3,
                h_external=h_external,
            )
        return residual

    zero = np.zeros(3)
    residual_zero = evaluate(zero)
    matrix = np.column_stack([evaluate(np.eye(3)[i]) - residual_zero for i in range(3)])
    condition_number = float(np.linalg.cond(matrix))
    if not np.isfinite(condition_number) or condition_number > 1e14:
        raise RuntimeError(f"Ill-conditioned slope system: cond={condition_number:.3e}")
    beta = np.linalg.solve(matrix, -residual_zero)
    residual, next_state = evaluate(beta, return_state=True)
    residual_norm = float(np.linalg.norm(residual, ord=np.inf))
    return next_state, {
        "beta": beta,
        "condition_number": condition_number,
        "linear_residual_inf": residual_norm,
        "pcm_specific_power_W_per_kg": float(pcm_specific_power),
        "layer2_average_C": float(layer2_average),
        "surface_start_C": float(surface_start),
        "h_external_W_per_m2K": float(h_external),
    }


def crossing_time(previous_time, current_time, previous_temperature, current_temperature, threshold):
    if previous_temperature > threshold >= current_temperature:
        fraction = (previous_temperature - threshold) / (
            previous_temperature - current_temperature
        )
        return previous_time + fraction * (current_time - previous_time)
    return None


def modal_snapshot(state, basis1, basis2, q_pcm, params):
    inner_surface = float(
        w1_values(0.0, state.body_temperature, state.interface_12, params)
        + state.b1 @ basis1.phi_left
    )
    outer_surface = layer3_surface(state, params)
    h_external = external_coefficient(outer_surface, params)
    layer2_average = modal_average_layer2(state, basis2, params)
    specific_power = q_pcm(layer2_average)
    body_power = params.body_area * params.h_body * (
        state.body_temperature - inner_surface
    )
    environment_power = params.body_area * h_external * (
        outer_surface - params.ambient_temperature
    )
    pcm_power = params.body_area * params.rho[1] * params.d[1] * specific_power
    return {
        "time_s": state.time_s,
        "time_h": state.time_s / 3600.0,
        "body_temperature_C": state.body_temperature,
        "inner_surface_C": inner_surface,
        "interface_12_C": state.interface_12,
        "interface_23_C": state.interface_23,
        "outer_surface_C": outer_surface,
        "layer2_average_C": layer2_average,
        "h_external_W_per_m2K": h_external,
        "pcm_specific_power_W_per_kg": specific_power,
        "body_to_clothing_power_W": body_power,
        "pcm_power_W": pcm_power,
        "environment_loss_W": environment_power,
    }


def modal_profile_rows(state, basis1, basis2, params, points_per_layer=21):
    rows = []
    offsets = np.cumsum(np.r_[0.0, params.d[:-1]])
    local_grids = [np.linspace(0.0, params.d[i], points_per_layer) for i in range(3)]
    temperatures = [
        w1_values(local_grids[0], state.body_temperature, state.interface_12, params)
        + evaluate_series(basis1, state.b1, local_grids[0]),
        w2_values(local_grids[1], state.interface_12, state.interface_23, params)
        + evaluate_series(basis2, state.b2, local_grids[1]),
        w3_values(local_grids[2], state.interface_23, state.h_external, params)
        + evaluate_series(state.basis3, state.b3, local_grids[2]),
    ]
    for layer in range(3):
        for y, temperature in zip(local_grids[layer], temperatures[layer]):
            rows.append(
                {
                    "time_s": state.time_s,
                    "time_h": state.time_s / 3600.0,
                    "layer": layer + 1,
                    "x_m": offsets[layer] + y,
                    "temperature_C": float(temperature),
                }
            )
    return rows


def simulate_modal(args, q_pcm, pcm_metadata, params=PARAMS):
    basis1 = make_layer1_basis(params, args.modes, args.quadrature)
    basis2 = make_layer2_basis(params, args.modes, args.quadrature)
    initial_h = external_coefficient(params.initial_temperature, params)
    basis3 = make_layer3_basis(params, initial_h, args.modes, args.quadrature)
    initial_layer3_residual = params.initial_temperature - w3_values(
        basis3.quad_y, params.initial_temperature, initial_h, params
    )
    state = ModalState(
        time_s=0.0,
        body_temperature=params.initial_temperature,
        interface_12=params.initial_temperature,
        interface_23=params.initial_temperature,
        b1=np.zeros(args.modes),
        b2=np.zeros(args.modes),
        b3=project_values(initial_layer3_residual, basis3),
        basis3=basis3,
        h_external=initial_h,
    )
    initial_energy = modal_total_energy(state, basis1, basis2, params)
    cumulative_pcm = 0.0
    cumulative_environment = 0.0
    t15 = None
    t10 = None
    max_condition = 0.0
    max_linear_residual = 0.0
    rows = [modal_snapshot(state, basis1, basis2, q_pcm, params)]
    profiles = modal_profile_rows(state, basis1, basis2, params)
    next_output = args.output_interval
    next_profile = args.profile_interval
    max_time_s = args.max_hours * 3600.0

    while state.time_s < max_time_s and state.body_temperature > 10.0:
        previous = state
        step = min(args.dt, max_time_s - state.time_s)
        state, diagnostic = modal_step(
            state,
            step,
            basis1,
            basis2,
            q_pcm,
            args.modes,
            args.quadrature,
            params,
        )
        max_condition = max(max_condition, diagnostic["condition_number"])
        max_linear_residual = max(max_linear_residual, diagnostic["linear_residual_inf"])
        snapshot_previous = modal_snapshot(previous, basis1, basis2, q_pcm, params)
        snapshot_current = modal_snapshot(state, basis1, basis2, q_pcm, params)
        cumulative_pcm += diagnostic["pcm_specific_power_W_per_kg"] * (
            params.body_area * params.rho[1] * params.d[1]
        ) * step
        cumulative_environment += 0.5 * (
            snapshot_previous["environment_loss_W"]
            + snapshot_current["environment_loss_W"]
        ) * step
        if t15 is None:
            t15 = crossing_time(
                previous.time_s,
                state.time_s,
                previous.body_temperature,
                state.body_temperature,
                15.0,
            )
        if t10 is None:
            t10 = crossing_time(
                previous.time_s,
                state.time_s,
                previous.body_temperature,
                state.body_temperature,
                10.0,
            )
        if state.time_s + 1e-9 >= next_output or state.body_temperature <= 10.0:
            rows.append(snapshot_current)
            while next_output <= state.time_s + 1e-9:
                next_output += args.output_interval
        if state.time_s + 1e-9 >= next_profile or state.body_temperature <= 10.0:
            profiles.extend(modal_profile_rows(state, basis1, basis2, params))
            while next_profile <= state.time_s + 1e-9:
                next_profile += args.profile_interval

    final_energy = modal_total_energy(state, basis1, basis2, params)
    energy_residual = (final_energy - initial_energy) - (
        cumulative_pcm - cumulative_environment
    )
    energy_scale = max(abs(cumulative_pcm) + abs(cumulative_environment), 1.0)
    summary = {
        "status": "PASS" if t10 is not None else "BUDGET_REACHED",
        "method": "modal",
        "model_id": MODEL_ID,
        "model_version": MODEL_VERSION,
        "dt_s": args.dt,
        "modes": args.modes,
        "quadrature_order": args.quadrature,
        "simulated_time_s": state.time_s,
        "simulated_time_h": state.time_s / 3600.0,
        "t15_s": t15,
        "t15_h": None if t15 is None else t15 / 3600.0,
        "t10_s": t10,
        "t10_h": None if t10 is None else t10 / 3600.0,
        "final_body_temperature_C": state.body_temperature,
        "maximum_slope_system_condition": max_condition,
        "maximum_linear_residual_inf": max_linear_residual,
        "cumulative_pcm_energy_J": cumulative_pcm,
        "cumulative_environment_loss_J": cumulative_environment,
        "global_energy_residual_J": energy_residual,
        "global_energy_relative_residual": energy_residual / energy_scale,
        "pcm_data": pcm_metadata,
    }
    return summary, rows, profiles


def build_fvm_grid(params: ModelParameters, cells_per_layer: int):
    positions = [0.0]
    edge_materials = []
    offset = 0.0
    for material in range(3):
        grid = np.linspace(0.0, params.d[material], cells_per_layer + 1)
        for value in grid[1:]:
            positions.append(offset + value)
            edge_materials.append(material)
        offset += params.d[material]
    positions = np.asarray(positions)
    edge_materials = np.asarray(edge_materials, dtype=int)
    edge_lengths = np.diff(positions)
    conductance = params.k[edge_materials] / edge_lengths
    capacity_area = np.zeros_like(positions)
    pcm_mass_area = np.zeros_like(positions)
    for edge, material in enumerate(edge_materials):
        heat_capacity = params.rho[material] * params.c[material] * edge_lengths[edge]
        capacity_area[edge] += 0.5 * heat_capacity
        capacity_area[edge + 1] += 0.5 * heat_capacity
        if material == 1:
            mass = params.rho[material] * edge_lengths[edge]
            pcm_mass_area[edge] += 0.5 * mass
            pcm_mass_area[edge + 1] += 0.5 * mass
    return positions, edge_materials, conductance, capacity_area, pcm_mass_area


def fvm_layer2_average(temperatures, positions, params):
    x1 = params.d[0]
    x2 = params.d[0] + params.d[1]
    mask = (positions >= x1 - 1e-15) & (positions <= x2 + 1e-15)
    return float(np.trapezoid(temperatures[mask], positions[mask]) / params.d[1])


def simulate_fvm(args, q_pcm, pcm_metadata, params=PARAMS):
    positions, edge_materials, conductance, capacity_area, pcm_mass_area = build_fvm_grid(
        params, args.fvm_cells
    )
    node_count = len(positions)
    initial = np.r_[
        params.initial_temperature,
        np.full(node_count, params.initial_temperature),
        0.0,
        0.0,
    ]

    def rhs(_time, values):
        body_temperature = values[0]
        temperatures = values[1 : 1 + node_count]
        layer_energy_rate = np.zeros(node_count)
        fluxes = conductance * (temperatures[:-1] - temperatures[1:])
        layer_energy_rate[:-1] -= fluxes
        layer_energy_rate[1:] += fluxes
        body_flux_area = params.h_body * (body_temperature - temperatures[0])
        layer_energy_rate[0] += body_flux_area
        h_external = external_coefficient(temperatures[-1], params)
        environment_flux_area = h_external * (
            temperatures[-1] - params.ambient_temperature
        )
        layer_energy_rate[-1] -= environment_flux_area
        average_temperature = fvm_layer2_average(temperatures, positions, params)
        specific_power = q_pcm(average_temperature)
        layer_energy_rate += pcm_mass_area * specific_power
        derivative = np.empty_like(values)
        derivative[0] = -params.body_area * body_flux_area / params.body_capacity
        derivative[1 : 1 + node_count] = layer_energy_rate / capacity_area
        derivative[-2] = params.body_area * np.sum(pcm_mass_area) * specific_power
        derivative[-1] = params.body_area * environment_flux_area
        return derivative

    def event15(_time, values):
        return values[0] - 15.0

    def event10(_time, values):
        return values[0] - 10.0

    event15.direction = -1
    event10.direction = -1
    event10.terminal = True
    solution = solve_ivp(
        rhs,
        (0.0, args.max_hours * 3600.0),
        initial,
        method="BDF",
        rtol=args.fvm_rtol,
        atol=args.fvm_atol,
        max_step=args.fvm_max_step,
        dense_output=True,
        events=(event15, event10),
    )
    if not solution.success:
        raise RuntimeError(solution.message)
    final_time = float(solution.t[-1])
    sample_times = np.arange(0.0, final_time, args.output_interval)
    if len(sample_times) == 0 or sample_times[-1] < final_time:
        sample_times = np.r_[sample_times, final_time]
    samples = solution.sol(sample_times)
    rows = []
    for index, time_s in enumerate(sample_times):
        body_temperature = float(samples[0, index])
        temperatures = samples[1 : 1 + node_count, index]
        average_temperature = fvm_layer2_average(temperatures, positions, params)
        specific_power = q_pcm(average_temperature)
        h_external = external_coefficient(float(temperatures[-1]), params)
        rows.append(
            {
                "time_s": float(time_s),
                "time_h": float(time_s / 3600.0),
                "body_temperature_C": body_temperature,
                "inner_surface_C": float(temperatures[0]),
                "interface_12_C": float(temperatures[args.fvm_cells]),
                "interface_23_C": float(temperatures[2 * args.fvm_cells]),
                "outer_surface_C": float(temperatures[-1]),
                "layer2_average_C": average_temperature,
                "h_external_W_per_m2K": h_external,
                "pcm_specific_power_W_per_kg": specific_power,
                "body_to_clothing_power_W": params.body_area
                * params.h_body
                * (body_temperature - temperatures[0]),
                "pcm_power_W": params.body_area
                * params.rho[1]
                * params.d[1]
                * specific_power,
                "environment_loss_W": params.body_area
                * h_external
                * (temperatures[-1] - params.ambient_temperature),
            }
        )
    profile_times = np.arange(0.0, final_time, args.profile_interval)
    if len(profile_times) == 0 or profile_times[-1] < final_time:
        profile_times = np.r_[profile_times, final_time]
    profile_values = solution.sol(profile_times)
    profiles = []
    boundaries = np.cumsum(params.d)
    for time_index, time_s in enumerate(profile_times):
        for node, x in enumerate(positions):
            layer = int(np.searchsorted(boundaries, x, side="left") + 1)
            layer = min(layer, 3)
            profiles.append(
                {
                    "time_s": float(time_s),
                    "time_h": float(time_s / 3600.0),
                    "layer": layer,
                    "x_m": float(x),
                    "temperature_C": float(profile_values[1 + node, time_index]),
                }
            )
    t15 = float(solution.t_events[0][0]) if len(solution.t_events[0]) else None
    t10 = float(solution.t_events[1][0]) if len(solution.t_events[1]) else None
    final_values = solution.y[:, -1]
    clothing_energy_initial = params.body_area * np.sum(
        capacity_area * params.initial_temperature
    )
    clothing_energy_final = params.body_area * np.sum(
        capacity_area * final_values[1 : 1 + node_count]
    )
    energy_change = (
        params.body_capacity * (final_values[0] - params.initial_temperature)
        + clothing_energy_final
        - clothing_energy_initial
    )
    cumulative_pcm = float(final_values[-2])
    cumulative_environment = float(final_values[-1])
    energy_residual = energy_change - (cumulative_pcm - cumulative_environment)
    energy_scale = max(abs(cumulative_pcm) + abs(cumulative_environment), 1.0)
    summary = {
        "status": "PASS" if t10 is not None else "BUDGET_REACHED",
        "method": "finite_volume",
        "model_id": MODEL_ID,
        "model_version": MODEL_VERSION,
        "cells_per_layer": args.fvm_cells,
        "node_count": node_count,
        "rtol": args.fvm_rtol,
        "atol": args.fvm_atol,
        "maximum_step_s": args.fvm_max_step,
        "solver_steps": int(len(solution.t)),
        "function_evaluations": int(solution.nfev),
        "simulated_time_s": final_time,
        "simulated_time_h": final_time / 3600.0,
        "t15_s": t15,
        "t15_h": None if t15 is None else t15 / 3600.0,
        "t10_s": t10,
        "t10_h": None if t10 is None else t10 / 3600.0,
        "final_body_temperature_C": float(final_values[0]),
        "cumulative_pcm_energy_J": cumulative_pcm,
        "cumulative_environment_loss_J": cumulative_environment,
        "global_energy_residual_J": energy_residual,
        "global_energy_relative_residual": energy_residual / energy_scale,
        "pcm_data": pcm_metadata,
    }
    return summary, rows, profiles


def write_csv(path: Path, rows):
    if not rows:
        raise ValueError(f"No rows for {path}")
    with path.open("w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def make_plots(output_dir: Path, rows, profiles, method: str):
    times = np.array([row["time_h"] for row in rows])
    plt.figure(figsize=(8.2, 5.0))
    for key, label in [
        ("body_temperature_C", "Body node"),
        ("inner_surface_C", "Inner fabric surface"),
        ("outer_surface_C", "Outer clothing surface"),
    ]:
        plt.plot(times, [row[key] for row in rows], label=label)
    plt.axhline(15.0, color="#C43C39", linestyle="--", linewidth=1.0, label="15 C threshold")
    plt.axhline(10.0, color="#7C2D2D", linestyle=":", linewidth=1.0, label="10 C threshold")
    plt.xlabel("Time (h)")
    plt.ylabel("Temperature (C)")
    plt.title(f"Q1 temperature history - {method}")
    plt.grid(alpha=0.25)
    plt.legend()
    plt.tight_layout()
    plt.savefig(output_dir / "temperature_history.png", dpi=180)
    plt.close()

    unique_times = sorted({row["time_s"] for row in profiles})
    selected = unique_times[-6:] if len(unique_times) > 6 else unique_times
    plt.figure(figsize=(8.2, 5.0))
    for time_s in selected:
        subset = [row for row in profiles if row["time_s"] == time_s]
        subset.sort(key=lambda row: row["x_m"])
        plt.plot(
            np.array([row["x_m"] for row in subset]) * 1000.0,
            [row["temperature_C"] for row in subset],
            label=f"{time_s / 3600.0:.2f} h",
        )
    plt.xlabel("Distance from body side (mm)")
    plt.ylabel("Temperature (C)")
    plt.title(f"Q1 clothing temperature profiles - {method}")
    plt.grid(alpha=0.25)
    plt.legend()
    plt.tight_layout()
    plt.savefig(output_dir / "temperature_profiles.png", dpi=180)
    plt.close()


def parse_args():
    repo_root = Path(__file__).resolve().parents[3]
    parser = argparse.ArgumentParser()
    parser.add_argument("--method", choices=["modal", "fvm"], required=True)
    parser.add_argument(
        "--pcm-file",
        type=Path,
        default=repo_root / "data" / "raw" / "附件1 放热能力数据.xlsx",
    )
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--max-hours", type=float, default=72.0)
    parser.add_argument("--output-interval", type=float, default=10.0)
    parser.add_argument("--profile-interval", type=float, default=300.0)
    parser.add_argument("--dt", type=float, default=1.0)
    parser.add_argument("--modes", type=int, default=30)
    parser.add_argument("--quadrature", type=int, default=128)
    parser.add_argument("--fvm-cells", type=int, default=24)
    parser.add_argument("--fvm-rtol", type=float, default=1e-7)
    parser.add_argument("--fvm-atol", type=float, default=1e-8)
    parser.add_argument("--fvm-max-step", type=float, default=2.0)
    return parser.parse_args()


def main():
    args = parse_args()
    if args.dt <= 0 or args.modes < 2 or args.quadrature < 2 * args.modes:
        raise SystemExit("Require dt>0, modes>=2, and quadrature>=2*modes")
    args.output_dir.mkdir(parents=True, exist_ok=True)
    q_pcm, pcm_metadata = load_pcm_curve(args.pcm_file)
    if args.method == "modal":
        summary, rows, profiles = simulate_modal(args, q_pcm, pcm_metadata)
    else:
        summary, rows, profiles = simulate_fvm(args, q_pcm, pcm_metadata)
    (args.output_dir / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    write_csv(args.output_dir / "body_temperature.csv", rows)
    write_csv(args.output_dir / "temperature_profiles.csv", profiles)
    make_plots(args.output_dir, rows, profiles, args.method)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if summary["status"] == "PASS" else 2


if __name__ == "__main__":
    raise SystemExit(main())
