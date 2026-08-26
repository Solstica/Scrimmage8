import csv
import json
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from scipy.integrate import cumulative_trapezoid


# 路径
root = Path(__file__).resolve().parents[3]
data_file = root / "data" / "raw" / "附件1 放热能力数据.xlsx"
out = root / "modules" / "20_q1" / "results" / "EXPERIMENT" / "dsc_analysis"
out.mkdir(parents=True, exist_ok=True)

# 读数据
ws = load_workbook(data_file, read_only=True, data_only=True)["Sheet1"]
data = np.array(list(ws.iter_rows(min_row=2, max_col=2, values_only=True)), dtype=float)
T = data[:, 0]
dsc = data[:, 1]
p_raw = -1000.0 * dsc

# 端点直线操作基线
baseline = np.interp(T, [T[0], T[-1]], [p_raw[0], p_raw[-1]])
p = p_raw - baseline
area_raw = np.trapezoid(p_raw, T)
area = np.trapezoid(p, T)

# 冷却方向累计，高温端 xi=0，低温端 xi=1
cum_raw = -cumulative_trapezoid(p_raw[::-1], T[::-1], initial=0.0)
cum = -cumulative_trapezoid(p[::-1], T[::-1], initial=0.0)
xi_raw = (cum_raw / cum_raw[-1])[::-1]
xi = (cum / cum[-1])[::-1]
w = p / area

peak = int(np.argmax(p))
half_index = np.flatnonzero(p >= 0.5 * p[peak])
T10 = float(np.interp(0.1, xi[::-1], T[::-1]))
T50 = float(np.interp(0.5, xi[::-1], T[::-1]))
T90 = float(np.interp(0.9, xi[::-1], T[::-1]))
T10_raw = float(np.interp(0.1, xi_raw[::-1], T[::-1]))
T50_raw = float(np.interp(0.5, xi_raw[::-1], T[::-1]))
T90_raw = float(np.interp(0.9, xi_raw[::-1], T[::-1]))

# 保存图表数据
with (out / "dsc_processed.csv").open("w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(
        [
            "temperature_C",
            "dsc_raw_mW_per_mg",
            "exothermic_power_raw_W_per_kg",
            "endpoint_linear_baseline_W_per_kg",
            "exothermic_power_corrected_W_per_kg",
            "solidification_progress_raw",
            "solidification_progress_corrected",
            "normalized_capacity_weight_per_C",
        ]
    )
    writer.writerows(zip(T, dsc, p_raw, baseline, p, xi_raw, xi, w))

summary = {
    "rows": int(len(T)),
    "temperature_range_C": [float(T[0]), float(T[-1])],
    "baseline_rule": "straight line through the two endpoints; operational baseline approved at model version 3a1982f",
    "endpoint_power_W_per_kg": [float(p_raw[0]), float(p_raw[-1])],
    "raw_peak_temperature_C": float(T[np.argmax(p_raw)]),
    "corrected_peak_temperature_C": float(T[peak]),
    "corrected_peak_power_W_per_kg": float(p[peak]),
    "corrected_half_peak_interval_C": [
        float(T[half_index[0]]),
        float(T[half_index[-1]]),
    ],
    "raw_area_W_K_per_kg": float(area_raw),
    "corrected_area_W_K_per_kg": float(area),
    "baseline_area_fraction": float(1.0 - area / area_raw),
    "progress_temperature_C": {"xi_0.1": T10, "xi_0.5": T50, "xi_0.9": T90},
    "raw_progress_temperature_C": {"xi_0.1": T10_raw, "xi_0.5": T50_raw, "xi_0.9": T90_raw},
    "progress_temperature_shift_C": {
        "xi_0.1": T10 - T10_raw,
        "xi_0.5": T50 - T50_raw,
        "xi_0.9": T90 - T90_raw,
    },
    "max_abs_progress_difference": float(np.max(np.abs(xi_raw - xi))),
    "latent_heat_relation": "L_pcm[J/kg] = corrected_area * 60 / scan_rate[K/min]",
    "latent_heat_numerator_J_K_per_kg_min": float(area * 60.0),
}
(out / "dsc_summary.json").write_text(
    json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
)

# 图1：原始曲线与基线候选
fig, ax = plt.subplots(figsize=(7.4, 4.6))
ax.plot(T, p_raw, color="#176B87", linewidth=2.0, label="Exothermic power (raw)")
ax.plot(T, baseline, color="#555555", linewidth=1.4, linestyle="--", label="Endpoint baseline")
ax.fill_between(T, baseline, p_raw, color="#64CCC5", alpha=0.28, label="Corrected peak area")
ax.scatter(T[peak], p_raw[peak], color="#C33C54", s=34, zorder=3)
ax.annotate(
    f"Peak {T[peak]:.3f} degC",
    (T[peak], p_raw[peak]),
    xytext=(14, -26),
    textcoords="offset points",
    fontsize=9,
)
ax.set_xlabel("Temperature (degC)")
ax.set_ylabel("Specific exothermic power (W/kg)")
ax.set_title("DSC exothermic curve and baseline candidate")
ax.grid(alpha=0.22)
ax.legend(frameon=False)
fig.tight_layout()
fig.savefig(out / "dsc_curve.png", dpi=200)
plt.close(fig)

# 图2：累计固化进度
fig, ax = plt.subplots(figsize=(7.4, 4.6))
ax.plot(T, xi, color="#0F766E", linewidth=2.2, label="Endpoint-baseline corrected")
ax.plot(T, xi_raw, color="#D97706", linewidth=1.4, linestyle="--", label="Without baseline correction")
for value, temp in [(0.1, T10), (0.5, T50), (0.9, T90)]:
    ax.scatter(temp, value, color="#C33C54", s=28, zorder=3)
    ax.annotate(f"{value:.0%}: {temp:.2f} C", (temp, value), xytext=(7, 5), textcoords="offset points", fontsize=8.5)
ax.set_xlabel("Temperature (degC)")
ax.set_ylabel("Solidification progress")
ax.set_title("Normalized solidification progress during cooling")
ax.set_ylim(-0.03, 1.03)
ax.invert_xaxis()
ax.grid(alpha=0.22)
ax.legend(frameon=False)
fig.tight_layout()
fig.savefig(out / "solidification_progress.png", dpi=200)
plt.close(fig)

# 图3：有效比热中的归一化潜热权重
fig, ax = plt.subplots(figsize=(7.4, 4.6))
ax.plot(T, w, color="#7A5195", linewidth=2.2)
ax.fill_between(T, 0.0, w, color="#BC8FCE", alpha=0.25)
ax.axvline(T[peak], color="#C33C54", linewidth=1.2, linestyle="--")
ax.set_xlabel("Temperature (degC)")
ax.set_ylabel("Normalized latent-capacity weight (1/degC)")
ax.set_title("Temperature distribution of finite latent capacity")
ax.grid(alpha=0.22)
fig.tight_layout()
fig.savefig(out / "latent_capacity_weight.png", dpi=200)
plt.close(fig)

print(json.dumps(summary, ensure_ascii=False, indent=2))
