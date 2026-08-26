import csv
import math
import sys
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from scipy.integrate import cumulative_trapezoid
from scipy.optimize import brentq


# 参数：m, kg, s, W, J, degC
d = np.array([0.0007, 0.0004, 0.0003])
rho = np.array([208.0, 552.3, 300.0])
c = np.array([4803.8, 2400.0, 5463.2])
k = np.array([0.068, 0.06, 0.0527])
T_b = 37.0
T_inf = -40.0
h_b = 3.0


def read_dsc(path):
    ws = load_workbook(path, read_only=True, data_only=True)["Sheet1"]
    data = np.array(list(ws.iter_rows(min_row=2, max_col=2, values_only=True)), dtype=float)
    T = data[:, 0]
    p_raw = -1000.0 * data[:, 1]
    base = np.interp(T, [T[0], T[-1]], [p_raw[0], p_raw[-1]])
    p = np.maximum(p_raw - base, 0.0)
    area = np.trapezoid(p, T)
    cum = -cumulative_trapezoid(p[::-1], T[::-1], initial=0.0)
    xi = (cum / cum[-1])[::-1]
    return T, p, xi, area


def get_xi(temp, T, xi):
    return float(np.interp(temp, T, xi, left=1.0, right=0.0))


def get_w(temp, T, p, area):
    return float(np.interp(temp, T, p / area, left=0.0, right=0.0))


def roots(length, conductivity, h, M):
    bi = h * length / conductivity
    z = np.empty(M)
    eps = 1e-10
    for m in range(1, M + 1):
        left = (m - 0.5) * math.pi + eps
        right = m * math.pi - eps
        z[m - 1] = brentq(lambda value: math.tan(value) + value / bi, left, right)
    return z / length


def gauss(length, n):
    z, w = np.polynomial.legendre.leggauss(n)
    return 0.5 * length * (z + 1.0), 0.5 * length * w


def finish_basis(length, lam, alpha, y, qw, phi, ends):
    return {
        "length": length,
        "lam": lam,
        "gamma": alpha * lam**2,
        "y": y,
        "qw": qw,
        "phi": phi,
        "norm": np.sum(phi**2 * qw, axis=1),
        "int0": phi @ qw,
        "int1": phi @ (qw * y),
        "phi_l": ends[0],
        "phi_r": ends[1],
        "dphi_l": ends[2],
        "dphi_r": ends[3],
    }


def basis1(M, nq):
    lam = roots(d[0], k[0], h_b, M)
    y, qw = gauss(d[0], nq)
    r = h_b / (k[0] * lam)
    phi = np.cos(lam[:, None] * y) + r[:, None] * np.sin(lam[:, None] * y)
    ends = (
        np.ones(M),
        np.cos(lam * d[0]) + r * np.sin(lam * d[0]),
        np.full(M, h_b / k[0]),
        -lam * np.sin(lam * d[0]) + h_b / k[0] * np.cos(lam * d[0]),
    )
    return finish_basis(d[0], lam, k[0] / (rho[0] * c[0]), y, qw, phi, ends)


def basis2(M, nq):
    lam = np.arange(1, M + 1) * math.pi / d[1]
    y, qw = gauss(d[1], nq)
    phi = np.sin(lam[:, None] * y)
    ends = (np.zeros(M), np.zeros(M), lam, lam * np.cos(lam * d[1]))
    return finish_basis(d[1], lam, 1.0, y, qw, phi, ends)


def basis3(h, M, nq):
    lam = roots(d[2], k[2], h, M)
    y, qw = gauss(d[2], nq)
    phi = np.sin(lam[:, None] * y)
    ends = (
        np.zeros(M),
        np.sin(lam * d[2]),
        lam,
        lam * np.cos(lam * d[2]),
    )
    return finish_basis(d[2], lam, k[2] / (rho[2] * c[2]), y, qw, phi, ends)


def project(values, b):
    return (b["phi"] @ (b["qw"] * values)) / b["norm"]


def eval1(b, coef, y):
    y = np.asarray(y)
    r = h_b / (k[0] * b["lam"])
    phi = np.cos(b["lam"][:, None] * y.ravel()) + r[:, None] * np.sin(
        b["lam"][:, None] * y.ravel()
    )
    return (coef @ phi).reshape(y.shape)


def eval_sin(b, coef, y):
    y = np.asarray(y)
    phi = np.sin(b["lam"][:, None] * y.ravel())
    return (coef @ phi).reshape(y.shape)


def w1(y, T12):
    den = k[0] + h_b * d[0]
    slope = h_b * (T12 - T_b) / den
    intercept = (k[0] * T12 + h_b * d[0] * T_b) / den
    return intercept + slope * np.asarray(y)


def w2(y, T12, T23):
    f = np.asarray(y) / d[1]
    return (1.0 - f) * T12 + f * T23


def w3(y, T23, h):
    slope = h * (T_inf - T23) / (k[2] + h * d[2])
    return T23 + slope * np.asarray(y)


def get_h(surface):
    return 2.38 * max(surface - T_inf, 0.0) ** 0.25


def layer2_avg(state, b2):
    return 0.5 * (state["T12"] + state["T23"]) + state["b2"] @ b2["int0"] / d[1]


def outer_surface(state):
    return float(w3(d[2], state["T23"], state["h3"]) + state["b3"] @ state["basis3"]["phi_r"])


def inner_surface(state, b1):
    return float(w1(0.0, state["T12"]) + state["b1"] @ b1["phi_l"])


def modal_step(state, dt, b1, b2, M, nq, T_dsc, p_dsc, xi_dsc, area, L_pcm):
    h = get_h(outer_surface(state))
    b3 = basis3(h, M, nq)
    y3 = b3["y"]
    old_T3 = w3(y3, state["T23"], state["h3"]) + eval_sin(
        state["basis3"], state["b3"], y3
    )
    b3_start = project(old_T3 - w3(y3, state["T23"], h), b3)

    T2_avg = layer2_avg(state, b2)
    xi_trial = get_xi(T2_avg, T_dsc, xi_dsc)
    advancing = xi_trial > state["xi"] + 1e-12
    c_eff = c[1] + L_pcm * get_w(T2_avg, T_dsc, p_dsc, area) if advancing else c[1]

    e1 = np.exp(-b1["gamma"] * dt)
    e2 = np.exp(-(k[1] / (rho[1] * c_eff)) * b2["lam"]**2 * dt)
    e3 = np.exp(-b3["gamma"] * dt)
    g1 = -np.expm1(-b1["gamma"] * dt) / b1["gamma"]
    gamma2 = (k[1] / (rho[1] * c_eff)) * b2["lam"]**2
    g2 = -np.expm1(-gamma2 * dt) / gamma2
    g3 = -np.expm1(-b3["gamma"] * dt) / b3["gamma"]

    p10, p11 = b1["int0"] / b1["norm"], b1["int1"] / b1["norm"]
    p20, p21 = b2["int0"] / b2["norm"], b2["int1"] / b2["norm"]
    p30, p31 = b3["int0"] / b3["norm"], b3["int1"] / b3["norm"]
    den1 = k[0] + h_b * d[0]
    den3 = k[2] + h * d[2]

    def advance(beta, full=False):
        beta12, beta23 = beta
        force1 = -(k[0] * beta12 / den1 * p10 + h_b * beta12 / den1 * p11)
        force2 = -beta12 * (p20 - p21 / d[1]) - beta23 * p21 / d[1]
        force3 = -beta23 * (p30 - h * p31 / den3)
        c1_end = state["b1"] * e1 + force1 * g1
        c2_end = state["b2"] * e2 + force2 * g2
        c3_end = b3_start * e3 + force3 * g3
        T12 = state["T12"] + beta12 * dt
        T23 = state["T23"] + beta23 * dt
        s1 = h_b * (T12 - T_b) / den1
        s2 = (T23 - T12) / d[1]
        s3 = h * (T_inf - T23) / den3
        r1 = k[0] * (s1 + c1_end @ b1["dphi_r"]) - k[1] * (
            s2 + c2_end @ b2["dphi_l"]
        )
        r2 = k[1] * (s2 + c2_end @ b2["dphi_r"]) - k[2] * (
            s3 + c3_end @ b3["dphi_l"]
        )
        if full:
            return np.array([r1, r2]), T12, T23, c1_end, c2_end, c3_end
        return np.array([r1, r2])

    r0 = advance(np.zeros(2))
    mat = np.column_stack([advance(np.eye(2)[j]) - r0 for j in range(2)])
    cond = float(np.linalg.cond(mat))
    beta = np.linalg.solve(mat, -r0)
    res, T12, T23, c1_end, c2_end, c3_end = advance(beta, True)
    new = {
        "t": state["t"] + dt,
        "T12": T12,
        "T23": T23,
        "b1": c1_end,
        "b2": c2_end,
        "b3": c3_end,
        "basis3": b3,
        "h3": h,
        "xi": max(state["xi"], xi_trial),
    }
    return new, cond, float(np.max(np.abs(res)))


def first_cross(t0, t1, T0, T1, limit):
    if T0 > limit >= T1:
        return t0 + (T0 - limit) / (T0 - T1) * (t1 - t0)
    return None


def snapshot(state, b1, b2, v, L_pcm):
    return [
        v,
        L_pcm / 1000.0,
        state["t"],
        inner_surface(state, b1),
        state["T12"],
        layer2_avg(state, b2),
        state["T23"],
        outer_surface(state),
        state["xi"],
    ]


def profile(state, b1, b2):
    rows = []
    ys = [np.linspace(0.0, d[i], 31) for i in range(3)]
    vals = [
        w1(ys[0], state["T12"]) + eval1(b1, state["b1"], ys[0]),
        w2(ys[1], state["T12"], state["T23"]) + eval_sin(b2, state["b2"], ys[1]),
        w3(ys[2], state["T23"], state["h3"]) + eval_sin(state["basis3"], state["b3"], ys[2]),
    ]
    offsets = np.r_[0.0, np.cumsum(d)[:-1]]
    for i in range(3):
        for y, temp in zip(ys[i], vals[i]):
            rows.append([state["t"], i + 1, offsets[i] + y, float(temp)])
    return rows


def solve_case(v, dt, M, nq, T_dsc, p_dsc, xi_dsc, area, save_detail=False):
    L_pcm = area * 60.0 / v
    b1 = basis1(M, nq)
    b2 = basis2(M, nq)
    h0 = get_h(T_b)
    b3 = basis3(h0, M, nq)
    b3_initial = project(np.full(nq, T_b) - w3(b3["y"], T_b, h0), b3)
    state = {
        "t": 0.0,
        "T12": T_b,
        "T23": T_b,
        "b1": np.zeros(M),
        "b2": np.zeros(M),
        "b3": b3_initial,
        "basis3": b3,
        "h3": h0,
        "xi": 0.0,
    }
    history = [snapshot(state, b1, b2, v, L_pcm)]
    profiles = profile(state, b1, b2) if save_detail else []
    t15 = None
    t10 = None
    max_cond = 0.0
    max_res = 0.0
    next_profile = 10.0

    while state["t"] < 1800.0 and t10 is None:
        old_t = state["t"]
        old_T = inner_surface(state, b1)
        state, cond, res = modal_step(
            state, dt, b1, b2, M, nq, T_dsc, p_dsc, xi_dsc, area, L_pcm
        )
        new_T = inner_surface(state, b1)
        max_cond = max(max_cond, cond)
        max_res = max(max_res, res)
        if t15 is None:
            t15 = first_cross(old_t, state["t"], old_T, new_T, 15.0)
        t10 = first_cross(old_t, state["t"], old_T, new_T, 10.0)
        history.append(snapshot(state, b1, b2, v, L_pcm))
        if save_detail and state["t"] + 1e-10 >= next_profile:
            profiles.extend(profile(state, b1, b2))
            next_profile += 10.0

    if t15 is None or t10 is None:
        raise RuntimeError(f"v={v:g} K/min did not reach both thresholds")
    if save_detail:
        profiles.extend(profile(state, b1, b2))
    return [v, L_pcm / 1000.0, t15, t10, max_cond, max_res, state["xi"]], history, profiles


def save_csv(path, header, rows):
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)


def main():
    root = Path(__file__).resolve().parents[3]
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else root / "modules" / "20_q1" / "results" / "formal"
    out.mkdir(parents=True, exist_ok=True)
    T_dsc, p_dsc, xi_dsc, area = read_dsc(root / "data" / "raw" / "附件1 放热能力数据.xlsx")

    results = []
    history = []
    profiles = []
    for v in [2.0, 5.0, 10.0]:
        ans, rows, prof = solve_case(
            v, 0.25, 40, 160, T_dsc, p_dsc, xi_dsc, area, save_detail=(v == 5.0)
        )
        results.append(ans)
        history.extend(rows)
        profiles.extend(prof)

    save_csv(
        out / "q1_results.csv",
        ["扫描速率（K/min）", "潜热（kJ/kg）", "t15（s）", "t10（s）", "最大条件数", "最大热流残差（W/m2）", "最终固化进度"],
        results,
    )
    save_csv(
        out / "q1_temperature_history.csv",
        ["扫描速率（K/min）", "潜热（kJ/kg）", "时间（s）", "贴身侧温度（℃）", "界面12温度（℃）", "功能层平均温度（℃）", "界面23温度（℃）", "外表面温度（℃）", "固化进度"],
        history,
    )
    save_csv(
        out / "q1_temperature_profiles.csv",
        ["时间（s）", "层号", "距人体侧位置（m）", "温度（℃）"],
        profiles,
    )
    for row in results:
        print(f"v={row[0]:g} K/min, L={row[1]:.3f} kJ/kg, t15={row[2]:.3f} s, t10={row[3]:.3f} s")


if __name__ == "__main__":
    main()
